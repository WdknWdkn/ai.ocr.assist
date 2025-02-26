#!/usr/bin/env python3
import sys
from parse_order_lambda import parse_excel
import openpyxl
from io import BytesIO
import logging

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Create a test Excel file
wb = openpyxl.Workbook()
ws = wb.active

# Add title row
ws.cell(row=1, column=1, value="会社別発注リスト（完工日：2024/12）")

# Add headers (row 2)
headers = ['業者ID', '業者名', 'コード', '建物名', '番号', '受付内容', '支払金額', '修繕作成者', '完工日', '修繕業者ID', '支払サイト', '支払日', '立替金', '請求日']
for col, header in enumerate(headers, 1):
    ws.cell(row=2, column=col, value=header)

# Add test data including rows that should be skipped
test_data = [
    # Valid data with all fields
    [1001, 'テスト業者1', '123-456', 'テストビル1', 1, '修繕内容1', 10000, '担当者A', '2024-12-01', 1234, 30, '2024-12-31', None, '2024-12-15'],
    # Row with vendor_id = 0 (should be skipped)
    [0, '', '789-012', '', 2, '', 0, '', '', None, None, '', None, ''],
    # Row with special date values and empty fields
    [1002, 'テスト業者2', '345-678', 'テストビル2', 3, '修繕内容2', None, '', '2999-12-31', None, None, '', None, '2999-12-31'],
    # Row with minimum required fields
    [1003, 'テスト業者3', '901-234', 'テストビル3', 4, '修繕内容3', 0, '', '', None, None, '', None, '']
]

for row_idx, row_data in enumerate(test_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Save to BytesIO
excel_bytes = BytesIO()
wb.save(excel_bytes)
excel_bytes.seek(0)

# Test parsing
try:
    result = parse_excel(excel_bytes.getvalue())
    logger.debug('Parsed result:')
    import json
    logger.debug(json.dumps(result, ensure_ascii=False, indent=2))
    
    # Verify results
    assert isinstance(result, dict), "Result should be a dictionary"
    assert "orders" in result, "Result should contain 'orders' key"
    assert "total_rows" in result, "Result should contain 'total_rows' key"
    assert "skipped_rows" in result, "Result should contain 'skipped_rows' key"
    assert "valid_rows" in result, "Result should contain 'valid_rows' key"
    
    assert result["total_rows"] == 4, "Should have 4 total rows"
    assert result["skipped_rows"] == 1, "Should have 1 skipped row"
    assert result["valid_rows"] == 3, "Should have 3 valid rows"
    assert len(result["orders"]) == 3, "Should have 3 valid orders"
    
    # Verify specific test cases
    orders = result["orders"]
    
    # Test case 1: All fields present
    assert orders[0]["業者ID"] == 1001, "First order should have vendor_id 1001"
    assert orders[0]["支払金額"] == 10000, "First order should have payment amount 10000"
    assert orders[0]["完工日"] == "2024-12-01", "First order should have completion date"
    
    # Test case 2: Special date values and empty fields
    assert orders[1]["業者ID"] == 1002, "Second order should have vendor_id 1002"
    assert orders[1]["支払金額"] == 0, "Second order should handle None payment as 0"
    assert orders[1]["完工日"] is None, "Second order should handle 2999 date as None"
    assert orders[1]["請求日"] is None, "Second order should handle 2999 date as None"
    
    # Test case 3: Minimum required fields
    assert orders[2]["業者ID"] == 1003, "Third order should have vendor_id 1003"
    assert orders[2]["支払金額"] == 0, "Third order should handle 0 payment"
    assert all(orders[2][field] is None for field in ["完工日", "支払日", "請求日"]), "Third order should have None dates"
    
    logger.debug("All assertions passed!")
except Exception as e:
    logger.error(f'Error: {str(e)}')
