#!/usr/bin/env python3
import json
import io
import os
import sys
import csv
import traceback
import openpyxl
from io import BytesIO

def parse_csv(file_bytes):
    """Parse CSV file content and return list of orders"""
    try:
        # Handle UTF-8 with BOM
        content = file_bytes.decode("utf-8", errors="ignore")
        if content.startswith('\ufeff'):
            content = content[1:]
            
        csv_stream = io.StringIO(content)
        reader = csv.DictReader(csv_stream)
        
        # Required fields check
        required_fields = ["業者ID", "業者名", "建物名", "番号", "受付内容", "支払金額", "完工日", "支払日", "請求日"]
        headers = [h.strip() for h in (reader.fieldnames or [])]
        
        # Check for missing required fields
        missing_fields = [field for field in required_fields if field not in headers]
        if missing_fields:
            raise ValueError(f"必須項目が見つかりません: {', '.join(missing_fields)}")
            
        orders = []
        for row_idx, row in enumerate(reader, start=2):  # start=2 because row 1 is headers
            # Skip empty rows
            if not any(row.values()):
                continue
                
            # Skip rows without vendor ID
            vendor_id = next((v.strip() for k, v in row.items() if k.strip() == "業者ID"), "")
            if not vendor_id:
                continue
                
            order = {}
            try:
                for field in required_fields:
                    # Find the matching header (accounting for whitespace)
                    header = next((k for k in row.keys() if k.strip() == field), None)
                    if header is None:
                        continue
                        
                    value = row[header].strip()
                    
                    # Numeric fields
                    if field in ["業者ID", "番号", "支払金額"]:
                        try:
                            order[field] = int(float(value)) if value else 0
                        except (ValueError, TypeError):
                            raise ValueError(f"{row_idx}行目の「{field}」の値が正しくありません")
                    # Date fields
                    elif field in ["完工日", "支払日", "請求日"]:
                        if not value:
                            raise ValueError(f"{row_idx}行目の「{field}」が入力されていません")
                        order[field] = value
                    # String fields
                    else:
                        if not value and field in ["業者名", "建物名", "受付内容"]:
                            raise ValueError(f"{row_idx}行目の「{field}」が入力されていません")
                        order[field] = value
                        
                orders.append(order)
            except ValueError as e:
                raise ValueError(str(e))
            except Exception as e:
                raise ValueError(f"{row_idx}行目のデータ処理中にエラーが発生しました: {str(e)}")
                
        if not orders:
            raise ValueError("有効なデータが見つかりません")
            
        return orders
        
    except ValueError as e:
        raise
    except Exception as e:
        print(f"Error parsing CSV: {str(e)}", file=sys.stderr)
        raise ValueError("CSVファイルの解析中にエラーが発生しました")

def parse_date(value):
    """Parse date value, handling special cases"""
    if not value:
        return None
    str_value = str(value).strip()
    # Skip placeholder dates
    if str_value.startswith("2999"):
        return None
    return str_value

def parse_excel(file_bytes):
    """Parse Excel file content and return list of orders"""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = wb.worksheets[0]  # 先頭シートを読む想定

        # タイトル行(1行目)をスキップし、2行目をヘッダー行として使用
        header_row = 2  # 2行目が必ずヘッダー行
        data_start_row = 3  # 3行目からデータ開始

        # 必須フィールドと任意フィールドの定義
        required_fields = ["業者ID", "業者名", "建物名", "番号", "受付内容"]
        optional_fields = ["支払金額", "完工日", "支払日", "請求日"]
        all_fields = required_fields + optional_fields
        
        # ヘッダー行から列インデックスを取得
        field_columns = {}
        # ヘッダー行の全列をチェック
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=header_row, column=col).value
            if header_value:
                header_str = str(header_value).strip()
                if header_str in all_fields:
                    field_columns[header_str] = col

        # 必須フィールドの存在チェック
        missing_fields = [field for field in all_fields if field not in field_columns]
        if missing_fields:
            raise ValueError(f"必須項目が見つかりません: {', '.join(missing_fields)}")

        # データ行を読み込み (3行目から)
        orders = []
        skipped_rows = 0
        for row_idx in range(data_start_row, sheet.max_row + 1):
            # 業者IDが空または0の行はスキップ
            vendor_id_cell = sheet.cell(row=row_idx, column=field_columns["業者ID"])
            if not vendor_id_cell.value or str(vendor_id_cell.value).strip() in ["", "0"]:
                skipped_rows += 1
                continue
                
            # ヘッダー行が繰り返される場合はスキップ
            if str(vendor_id_cell.value).strip() == "業者ID":
                continue

            order = {}
            row_has_data = False
            try:
                # 必須フィールドの処理
                for field in required_fields:
                    value = sheet.cell(row=row_idx, column=field_columns[field]).value
                    if value is not None:
                        row_has_data = True
                        str_value = str(value).strip()
                        
                        # 業者IDの処理
                        if field == "業者ID":
                            try:
                                vendor_id = int(float(str_value)) if str_value else 0
                                if vendor_id == 0:
                                    raise ValueError("業者IDが0または空です")
                                order[field] = vendor_id
                            except (ValueError, TypeError):
                                raise ValueError(f"{row_idx}行目の「{field}」の値が正しくありません")
                        # 番号の処理
                        elif field == "番号":
                            try:
                                order[field] = int(float(str_value)) if str_value else 0
                            except (ValueError, TypeError):
                                raise ValueError(f"{row_idx}行目の「{field}」の値が正しくありません")
                        # 文字列フィールドの処理
                        else:
                            if not str_value:
                                raise ValueError(f"{row_idx}行目の「{field}」が入力されていません")
                            order[field] = str_value
                            
                # 任意フィールドの処理
                for field in optional_fields:
                    value = sheet.cell(row=row_idx, column=field_columns[field]).value
                    if value is not None:
                        str_value = str(value).strip()
                        
                        # 支払金額の処理
                        if field == "支払金額":
                            try:
                                order[field] = int(float(str_value)) if str_value else 0
                            except (ValueError, TypeError):
                                order[field] = 0
                        # 日付フィールドの処理
                        else:
                            order[field] = parse_date(value)
                
                # 行にデータがあり、必須フィールドが揃っている場合のみ追加
                if row_has_data and all(field in order for field in required_fields):

                    orders.append(order)
            except ValueError as e:
                raise ValueError(str(e))
            except Exception as e:
                raise ValueError(f"{row_idx}行目のデータ処理中にエラーが発生しました: {str(e)}")

        if not orders:
            raise ValueError("有効なデータが見つかりません。3行目以降にデータが存在することを確認してください。")

        return {
            "orders": orders,
            "total_rows": sheet.max_row - data_start_row + 1,
            "skipped_rows": skipped_rows,
            "valid_rows": len(orders)
        }

    except ValueError as e:
        raise
    except Exception as e:
        print(f"Error parsing Excel: {str(e)}", file=sys.stderr)
        raise ValueError("Excelファイルの解析中にエラーが発生しました")

if __name__ == "__main__":
    print("This module is now used as a library and should not be run directly.")
